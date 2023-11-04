# Importing Libs
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import (PieChart, ProjectedPieChart, Reference)
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl import Workbook
from datetime import datetime
from os import environ
from sqs import SQS
import openpyxl
import logging
import boto3
import json
import ssm
import S3

# Configure Services
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
sqs_handler= SQS(logger)
ews_sqs_name = "vdf-dev-01-sqs-rca-send-emails"
# ssm = boto3.client('ssm', region_name='eu-central-1')
aws_region = environ["AWS_REGION"]
ssm_handler = ssm.SSM(logger, aws_region)
Mail = environ["DEVOPS_EMAIL"]


str_quary= "select j.name as 'JobName',run_date,message From msdb.dbo.sysjobs j INNER JOIN msdb.dbo.sysjobhistory h ON j.job_id = h.job_id where j.enabled = 1   and (message like '%The job succeeded%' or message like '%The job failed%') and  CONVERT(datetime, CONVERT(varchar(8), run_date), 112) BETWEEN GETDATE()-7 AND GETDATE()order by run_date desc"
str_tablescount_quary ="SELECT QUOTENAME(SCHEMA_NAME(sOBJ.schema_id)) + '.' + QUOTENAME(sOBJ.name) AS [TableName], SUM(sPTN.Rows) AS [RowCount] FROM sys.objects AS sOBJ INNER JOIN sys.partitions AS sPTN ON sOBJ.object_id = sPTN.object_id WHERE sOBJ.type = 'U' AND sOBJ.is_ms_shipped = 0x0 AND index_id < 2 GROUP BY sOBJ.schema_id , sOBJ.name ORDER BY [RowCount] desc "


DB_username = ssm_handler.get_parameter(environ["DB_USER"])
DB_password = ssm_handler.get_parameter(environ["DB_PASS"])
DB_server = environ["DB_SERVER"]

elk = "https://kibana.iaprobotics-dev.vodafone.com"
api = "/api/account/authenticate"
headers = {"Content-type": "application/json"}
odata_lic = "/odata/Settings/UiPath.Server.Configuration.OData.GetLicense"
odata_machines = "/odata/Machines"
orchURL = "https://orchestrator.iaprobotics-dev.vodafone.com"
auth = {
    "tenancyName":"Default",
    "usernameOrEmailAddress":environ["ORCH_AUTH_UN"],
    "password":environ["ORCH_AUTH_PW"]
}

redmine = "redmine.iaprobotics-dev.vodafone.com"
gitlab = "gitlab.iaprobotics-dev.vodafone.com"
kibana = "kibana.iaprobotics-dev.vodafone.com"
ELKApi = "YVhjUFpJa0J5OThqbzloekNodVU6bWhVYlFPd0NRbWUzWjdiaXRuU0k3UQ=="
ELKApiJson = {
    "id":"aXcPZIkBy98jo9hzChuU",
    "name":"Auditing",
    "api_key":"mhUbQOwCQme3Z7bitnSI7Q",
    "encoded":"YVhjUFpJa0J5OThqbzloekNodVU6bWhVYlFPd0NRbWUzWjdiaXRuU0k3UQ=="}

orchestrator = "orchestrator.iaprobotics-dev.vodafone.com"
OAppID = "a489cf6a-4827-4e77-bf5d-828c711623ee"
OAppSecret = "V7Mqt6gpihqf^h8d"

grafana = "grafana.iaprobotics-dev.vodafone.com"

Dev_ids = {
    # win ec2
    "Orchestrator A-V22": "i-07acfaaa448c15319",
    "Orchestrator B-V22": "i-083d678a3bf081c76",
    #"Orchestrator C-V22": "i-05019f5ec7fb057f5",

    # linux ec2
    "Orchestrator HAA A-V22": "i-0222595af9c4a173b",
    "Orchestrator HAA B-V22": "i-0e5dc0971254cc14b",
    "Orchestrator HAA C-V22": "i-050aecac94555d7d6",   
    "ElasticSearch Data A": "i-01ec11e4a4c9f8906",
    "ElasticSearch Data B": "i-0f7e8886046099a30",
    "ElasticSearch Data C": "i-0dd254050e17fdaaa",
    "ElasticSearch Master A": "i-0678c70bd5166e0dd",
    "ElasticSearch Master B": "i-01aa435ee858da7df",
    "ElasticSearch Master C": "i-00d6167c9be3c4eb2",
    "ElasticSearch Italy Data A": "i-0118afc4cbbc759a7",
    "ElasticSearch Italy Data B": "i-04674322c83e733fa",
    "ElasticSearch Italy Data C": "i-0c682afae28353682",
    "ElasticSearch Italy Master A": "i-0169c07acfdfaabe9",
    "ElasticSearch Italy Master B": "i-0253b1bf7ff45d5c8",
    "ElasticSearch Italy Master C": "i-0c6fd2ac7aae72f56",
    "Kibana A": "i-093ed8d4f56615c0c",
    "Kibana B": "i-043f9b4347435cabd",
    # "Kibana Italy A": "i-0e17ddf66a771bfad",
    # "Kibana Italy B": "i-0ba1d4d62b6f231fc",
    "FSX Server": "i-043f9b4347435cabd",
    "Redmine": "i-00f2420febd145ad5",
    "Gitlab": "i-01e6c7a9d5f53edad",
    "Prometheus": "i-095f89aadb7a58e3d",
    "Grafana": "i-0b9d52d3a505632a4",
    "Logstash A": "i-042b62fcfe82cbe2b",
    "Logstash B": "i-04c594fcb7690639f"
}

FSX = {
        "vdf-dev-01-fsx": "fs-05e8b3000bf708d75",
        "vss-robotics-DEV-fsx": "fs-085d949cce8d0c9d6",
        "vss-robotics-DEV-fsx-iap-fs": "fs-0d0f3674f9360bfd1"
    }

RDS = [
        # "vdf-dev-01-rds-gitlab-postgres",
        # "vdf-dev-01-rds-handset",
        "vdf-dev-01-rds-redmine-postgres",
        # "vdf-dev-01-rds-sql",
        "vdf-dev-01-rds-sql-v22"
        ]


DataBaseRec = [
    'QueueItems',
    'QueueItemsEvents',
    'QueueItemsComments',
    'Logs',
    'UserNotifications',
    'TenantNotifications',
    'Jobs',
    'AuditLogs',
    'AuditLogEntities',
    'Tasks',
    'RobotLicenseLogs',
    'RobotLicense'
    ]

# Time formating
today = datetime.now()
month = today.strftime("%b")
now = datetime.utcnow()

# file paths
file_path = '/tmp/RPA_Audit.xlsx'
file_chart = '/tmp/pie.xlsx'

# create excel file/sheets
def create_excel_file(sheet_names, file_path):
    wb = openpyxl.Workbook()
    for sheet_name in sheet_names:
        wb.create_sheet(sheet_name)
    std = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)
    wb.save(file_path)

# write to excel sheet
def write_to_excel_sheet(sheet_name, data, toprow):
    # S3.download_file('audit-iap-automation', 'RPA_Audit.xlsx', '/tmp/RPA_Audit.xlsx')
    # file_path = '/tmp/RPA_Audit.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    col = 1
    for Rvalue in toprow:
        sheet.cell(row=1, column=col).value = Rvalue
        col = col + 1
    row = sheet.max_row + 1
    for i, item in enumerate(data):
        sheet.cell(row=row, column=i+1).value = item
    wb.save(file_path)

def write_to_excel_sheet_custom(sheet_name, data, toprow, column_num_of_data):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    col = column_num_of_data
    sheet.cell(row=1, column=col).value = toprow
    roww = 2
    for i, item in enumerate(data):
        sheet.cell(row=roww, column=col).value = item
        roww = roww + 1
        # print(item)
    wb.save(file_path)

# upload file to S3
def uploadfile():
    fileN = 'DEV-RPA_Audit_for(' + str(month) + ').xlsx'
    S3.upload_file(file_path, 'audit-iap-automation', fileN)

# Sending Mail with Confirmation
    str_html = f"""
              Robotics IAP Housekeeping and Governance Checklist [DEV] \n
              Auditing Sheet Created under S3  \n
              Bucket Name: audit-iap-automation \n
              Link: https://audit-iap-automation.s3.eu-central-1.amazonaws.com \n
              \n
              Regards \n
              RPA - DevOps Team"""
              
    mail_message = {
        f"IssueID": str(int(now.timestamp())),
        "body":str_html,
        "subject":"Hi All, This is the Automated Audit file for this Month",
        "to": [Mail],
        "HasAttachment":"false"
    }
    
    #,"mohamed.elshenawy1@vodafone.com","ghada.ibrahim1@vodafone.com"],
    #, "fileName":"DB_Tables_Counts.CSV"}
    
    sqs_handler.send_message(ews_sqs_name,json.dumps(mail_message))
    
# Adjust Column Width
def column_width(sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    sheet = wb[sheet_name]
    column_widths = []
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width                    
    wb.save(file_path)

# Style Cells
def style_cells(sheet_name, md=50, hi=70):
    thresh_md = md
    thresh_hi = hi
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    gray_fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    sheet = wb[sheet_name]
    data = sheet.iter_rows()
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    flag = 0
    font = Font(name='Calibri',size=14,bold=True)
    # Iterate over cells
    for row in data:
        for cell in row:
            if(flag == 0):
                cell.fill = gray_fill
                cell.font = font
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            if isinstance(cell.value, float) or isinstance(cell.value, int):
                # Style borders
                # Style colors
                if cell.value < thresh_md:
                    cell.fill = green_fill
                elif thresh_md <= cell.value < thresh_hi:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
        flag = flag + 1
    # ws.page_setup.fitToHeight = 0
    # ws.page_setup.fitToWidth = 1
    wb.save(file_path)

# draw charts
def create_chart(Src_sheetN, src_col, max_row):
# Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
# Select the sheet you want to loop through
    sheet = wb[Src_sheetN]
    title = ""
# Configure Chart
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = Src_sheetN
    chart1.y_axis.title = 'Utilization Percentage'
    chart1.x_axis.title = 'Server Name'
# configure Labels
    labels = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    data = Reference(sheet, min_col=src_col, min_row=2, max_row=max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(labels)
    chart1.shape = 4
# Add and Save Changes
    sheet.add_chart(chart1, "H3")
    wb.save(file_path)